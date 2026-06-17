#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script para extrair texto de PDF e exibir linha por linha na tela
"""

import csv
import re
import sys
import os
from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from extrato_util import eh_descricao_saldo

try:
    import pdfplumber
except ImportError:
    print("Erro: pdfplumber não encontrado. Instale com: pip install pdfplumber")
    sys.exit(1)


def extrair_texto_pdf(caminho_pdf, debug=False):
    """
    Extrai texto de todas as páginas do PDF usando pdfplumber.
    Retorna (texto_com_pipes, linhas) para permitir detecção de formato.
    """
    linhas = []
    try:
        with pdfplumber.open(caminho_pdf) as pdf:
            if debug:
                print(f"PDF possui {len(pdf.pages)} páginas")
            for numero_pagina, page in enumerate(pdf.pages, 1):
                text = page.extract_text()
                if text:
                    pag_linhas = [ln.strip() for ln in text.split('\n') if ln.strip()]
                    if debug:
                        print(f"\n--- PÁGINA {numero_pagina} ---")
                        for i, linha in enumerate(pag_linhas, 1):
                            print(f"Linha {i}: {linha}")
                    linhas.extend(pag_linhas)
        texto = '|'.join(linhas) if linhas else None
        return texto, linhas
    except FileNotFoundError:
        print(f"Erro: Arquivo '{caminho_pdf}' não encontrado.")
        return None, []
    except Exception as e:
        print(f"Erro ao processar o PDF: {e}")
        return None, []


def extrair_periodo_sicoob(linhas):
    """
    Extrai o período do cabeçalho do extrato Sicoob.
    Ex.: PERÍODO: 01/01/2026 - 31/01/2026
    """
    texto_cabecalho = '\n'.join(linhas[:40])
    match = re.search(
        r'PER[IÍ]ODO:\s*(\d{2}/\d{2}/\d{4})\s*-\s*(\d{2}/\d{2}/\d{4})',
        texto_cabecalho,
        re.IGNORECASE,
    )
    if not match:
        return None, None, None

    inicio = match.group(1)
    fim = match.group(2)
    ano = inicio.split('/')[-1]
    return inicio, fim, ano


def extrair_dados_conta_sicoob(linhas):
    """
    Extrai cooperativa e conta do cabeçalho do extrato Sicoob.

    Exemplo no PDF:
      COOP.: 3067-8 / SICOOB - CREDIAUC/SC
      CONTA: 53.513-3 / LIMARIN VEICULOS LTDA.
    """
    texto_cabecalho = '\n'.join(linhas[:40])

    cooperativa = ''
    numero_conta = ''
    titular = ''

    match_coop = re.search(r'COOP\.?:\s*([\d\-]+)', texto_cabecalho, re.IGNORECASE)
    if match_coop:
        cooperativa = match_coop.group(1).strip()

    match_conta = re.search(
        r'CONTA:\s*([\d.\-]+)\s*/\s*(.+)',
        texto_cabecalho,
        re.IGNORECASE,
    )
    if match_conta:
        numero_conta = match_conta.group(1).strip()
        titular = match_conta.group(2).strip()

    # Padrão OFX Sicoob: BRANCHID=3067-8, ACCTID=53513-3
    acct_id = numero_conta.replace('.', '') if numero_conta else '00000000'

    return {
        'cooperativa': cooperativa,
        'branch_id': cooperativa,
        'numero_conta': numero_conta,
        'titular': titular,
        'acct_id': acct_id,
    }


def detectar_formato_sicoob(texto_ou_linhas):
    """
    Detecta o layout do extrato SICOOB pelo cabeçalho:
    - '4col': DATA | DOCUMENTO | HISTÓRICO | VALOR (limarin)
    - '3col': DATA | HISTÓRICO | VALOR (fabiane)
    """
    if isinstance(texto_ou_linhas, str):
        trecho = texto_ou_linhas[:3000].upper()
    else:
        trecho = ' '.join(str(x) for x in texto_ou_linhas[:80]).upper()
    # Normalizar acentos para comparação
    trecho_norm = trecho.replace('Ó', 'O').replace('Í', 'I')
    if 'DOCUMENTO' in trecho_norm and 'HISTORICO' in trecho_norm:
        return '4col'
    return '3col'

def organizar_lancamentos_por_data(texto):
    """
    Organiza os lançamentos em uma estrutura de dados baseada na data (dd/mm)
    
    Args:
        texto (str): Texto extraído do PDF
        
    Returns:
        dict: Dicionário com lançamentos organizados por data
    """
    # Padrão para identificar datas no formato dd/mm
    padrao_data = r'^\d{2}/\d{2}'
    
    # Dividir o texto em linhas e limpar
    linhas = texto.split('|')
    linhas_limpas = []
    
    # Limpar e processar linhas
    for linha in linhas:
        linha = linha.strip()
        if not linha:  # Pular linhas vazias
            continue
        
        # Verificar se a linha contém múltiplas datas (problema de concatenação)
        datas_encontradas = re.findall(r'\d{2}/\d{2}', linha)
        if len(datas_encontradas) > 1:
            # Dividir a linha em múltiplas partes baseado nas datas
            partes = re.split(r'(\d{2}/\d{2})', linha)
            for i in range(1, len(partes), 2):  # Pular o primeiro elemento vazio
                if i + 1 < len(partes):
                    nova_linha = partes[i] + partes[i + 1]
                    if nova_linha.strip():
                        linhas_limpas.append(nova_linha.strip())
        else:
            # Verificar se há uma data no meio da linha (problema de concatenação)
            match_data_meio = re.search(r'(\d{2}/\d{2})', linha)
            if match_data_meio and not linha.startswith(match_data_meio.group(1)):
                # Encontrar a posição da data no meio da linha
                pos_data = linha.find(match_data_meio.group(1))
                if pos_data > 0:
                    # Separar em duas linhas
                    linha_antes = linha[:pos_data].strip()
                    linha_depois = linha[pos_data:].strip()
                    if linha_antes:
                        linhas_limpas.append(linha_antes)
                    if linha_depois:
                        linhas_limpas.append(linha_depois)
                else:
                    linhas_limpas.append(linha)
            else:
                linhas_limpas.append(linha)
    
    # Dicionário para armazenar os lançamentos por data
    lancamentos_por_data = {}
    lancamento_atual = []
    data_atual = None
    valor_pendente = None
    padrao_valor_isolado = r'^\d{1,3}(?:\.\d{3})*,\d{2}$'

    for linha in linhas_limpas:
        linha = linha.strip()
        if re.match(padrao_valor_isolado, linha):
            valor_pendente = linha
            continue

        # Verificar se a linha começa com uma data (dd/mm)
        if re.match(padrao_data, linha):
            # Se já temos um lançamento em andamento, salvá-lo
            if data_atual and lancamento_atual:
                if data_atual not in lancamentos_por_data:
                    lancamentos_por_data[data_atual] = []
                # Juntar as linhas do lançamento em uma única string
                lancamento_completo = ' | '.join(lancamento_atual)
                lancamentos_por_data[data_atual].append(lancamento_completo)

            # Iniciar novo lançamento
            data_atual = linha[:5]  # Pegar apenas dd/mm
            lancamento_atual = [linha]
            # Valor isolado pertence ao lançamento apenas se a linha da data
            # não trouxer valor embutido (ex.: 02/01 DÉB.TIT... seguido de 13.540,50).
            if valor_pendente and not re.search(r'\d{1,3}(?:\.\d{3})*,\d{2}[DC]?', linha):
                lancamento_atual.append(valor_pendente)
            valor_pendente = None
        else:
            # Adicionar linha ao lançamento atual
            if data_atual:
                lancamento_atual.append(linha)
    
    # Adicionar o último lançamento
    if data_atual and lancamento_atual:
        if data_atual not in lancamentos_por_data:
            lancamentos_por_data[data_atual] = []
        # Juntar as linhas do lançamento em uma única string
        lancamento_completo = ' '.join(lancamento_atual)
        lancamentos_por_data[data_atual].append(lancamento_completo)
    
    return lancamentos_por_data

def exibir_lancamentos_organizados(lancamentos_por_data, debug=False):
    """
    Exibe os lançamentos organizados por data
    
    Args:
        lancamentos_por_data (dict): Dicionário com lançamentos organizados
        debug (bool): Se True, exibe a lista de lançamentos
    """
    if debug:
        print("\n" + "="*60)
        print("LISTA DE LANÇAMENTOS")
        print("="*60)
        
        # Ordenar as datas
        datas_ordenadas = sorted(lancamentos_por_data.keys())
        
        for data in datas_ordenadas:
            lancamentos = lancamentos_por_data[data]
            for lancamento in lancamentos:
                print(lancamento)
        
        print(f"\n📊 Total de datas: {len(datas_ordenadas)}")
        total_lancamentos = sum(len(lancamentos) for lancamentos in lancamentos_por_data.values())
        print(f"📊 Total de lançamentos: {total_lancamentos}")

def extrair_data_valor(lancamento, formato='3col', ano_referencia=None):
    """
    Extrai data, valor e tipo de um lançamento.
    4col: DATA DOCUMENTO HISTÓRICO VALOR - data dd/mm/yyyy, valor com D/C (450,00D)
    3col: DATA HISTÓRICO VALOR - data dd/mm, valor e C/D podem estar separados
    """
    partes = lancamento.split('|')
    primeira_parte = partes[0].strip()
    segunda_parte = partes[1].strip() if len(partes) > 1 else ""

    # Data: dd/mm/yyyy ou dd/mm
    match_full = re.match(r'^(\d{2}/\d{2}/\d{4})', primeira_parte)
    match_short = re.match(r'^(\d{2}/\d{2})(?!/)', primeira_parte)
    if match_full:
        data_processada = match_full.group(1)
    elif match_short:
        ano = ano_referencia or str(datetime.now().year)
        data_processada = f"{match_short.group(1)}/{ano}"
    else:
        data_processada = None

    # Valor: com D/C concatenado (4col: 450,00D) ou separado (3col: | C |)
    padrao_valor = r'(\d{1,3}(?:\.\d{3})*,\d{2})(?:\s*\|\s*([DC])|([DC]))'
    match_valor = re.search(padrao_valor, primeira_parte + "|" + segunda_parte)

    valor_processado = None
    tipo_operacao = None
    
    if match_valor:
        valor_str = match_valor.group(1)
        tipo_concat = match_valor.group(2) or match_valor.group(3) or None
        valor_limpo = valor_str.replace('.', '').replace(',', '.')
        try:
            valor_float = float(valor_limpo)

            # Buscar tipo: concatenado ou na segunda parte
            tipo = tipo_concat
            if not tipo and len(partes) > 1:
                # Procurar D ou C na segunda parte
                segunda_parte_tipo = partes[1].strip()
                if segunda_parte_tipo in ['D', 'C']:
                    tipo = segunda_parte_tipo

            if tipo == 'D':
                valor_float = -valor_float
            elif tipo == 'C':
                valor_float = abs(valor_float)
            else:
                tipo = None

            valor_processado = valor_float
            tipo_operacao = tipo
        except ValueError:
            pass
    elif len(partes) > 2:
        valor_isolado = partes[1].strip()
        tipo_separado = partes[2].strip()
        if re.match(r'^\d{1,3}(?:\.\d{3})*,\d{2}$', valor_isolado) and tipo_separado in ('D', 'C'):
            try:
                valor_float = float(valor_isolado.replace('.', '').replace(',', '.'))
                if tipo_separado == 'D':
                    valor_float = -valor_float
                valor_processado = valor_float
                tipo_operacao = tipo_separado
            except ValueError:
                pass

    return data_processada, valor_processado, tipo_operacao

def extrair_cnpj_cpf(texto):
    """
    Extrai o primeiro CNPJ ou CPF encontrado em um texto.
    Retorna string vazia se não encontrar.
    Aceita CNPJ com barra ou espaço e CPF com asteriscos.
    Busca apenas a partir da segunda parte da linha (ignora data/valor).
    """
    import re
    
    # Dividir a linha pelos separadores | para buscar em cada parte
    partes = texto.split('|')
    
    # Padrão CNPJ: 00.000.000/0000-00 ou 00.000.000 0000-00
    padrao_cnpj = r'\d{2}\.\d{3}\.\d{3}[ /]\d{4}-\d{2}'
    # Padrão CPF: 000.000.000-00 ou ***.000.000-00 ou ***.000.000-000 ou ***.000.000-**
    padrao_cpf = r'(?:\*{3}|\d{3})\.\d{3}\.\d{3}-(?:\d{2,3}|\*{2})'
    
    # Buscar a partir da segunda parte (ignorar primeira parte com data/valor)
    for i in range(1, len(partes)):
        parte = partes[i].strip()
        # Buscar CNPJ primeiro
        match_cnpj = re.search(padrao_cnpj, parte)
        if match_cnpj:
            return match_cnpj.group(0)
        # Buscar CPF
        match_cpf = re.search(padrao_cpf, parte)
        if match_cpf:
            return match_cpf.group(0)
    
    return ''

def extrair_saldo_final_resumo(lancamento):
    """
    Extrai o saldo final do resumo que aparece no formato "SALDO EM C.CORRENTE(+): 9.185,57C"
    """
    import re
    
    # Padrão para saldo em conta corrente
    padrao_saldo = r'SALDO EM C\.CORRENTE\(\+\):\s*(\d{1,3}(?:\.\d{3})*,\d{2})'
    match = re.search(padrao_saldo, lancamento)
    if match:
        valor_str = match.group(1)
        valor_limpo = valor_str.replace('.', '').replace(',', '.')
        try:
            return float(valor_limpo)
        except ValueError:
            pass
    
    # Padrão para saldo disponível
    padrao_disponivel = r'SALDO DISPONÍVEL\(=\):\s*(\d{1,3}(?:\.\d{3})*,\d{2})'
    match = re.search(padrao_disponivel, lancamento)
    if match:
        valor_str = match.group(1)
        valor_limpo = valor_str.replace('.', '').replace(',', '.')
        try:
            return float(valor_limpo)
        except ValueError:
            pass
    
    return None

def extrair_documento_historico_4col(lancamento_completo):
    """
    Layout 4col: DATA | DOCUMENTO | HISTÓRICO | VALOR.
    O HISTÓRICO pode ter várias linhas (ex: DÉB.TRANSF..., INTERCREDIS, FAV.: ..., etc.).
    Inclui todas as partes do lancamento na coluna historico.
    """
    partes = [p.strip() for p in lancamento_completo.strip().split('|') if p.strip()]
    primeira = partes[0] if partes else ''

    restante = re.sub(r'^\d{2}/\d{2}(/\d{4})?\s+', '', primeira)
    restante = re.sub(r'\s+\d{1,3}(?:\.\d{3})*,\d{2}[DC]?\s*$', '', restante).strip()
    tokens = restante.split()
    if not tokens:
        documento = ''
        historico_primeira = ''
    elif len(tokens) == 1:
        documento = tokens[0]
        historico_primeira = ''
    else:
        documento = tokens[0]
        historico_primeira = ' '.join(tokens[1:])

    # Juntar historico da primeira linha + todas as linhas de continuação
    continuacoes = partes[1:] if len(partes) > 1 else []
    historico_full = historico_primeira
    for c in continuacoes:
        if c and c not in ('C', 'D'):
            historico_full = (historico_full + ' ' + c).strip()
    return documento, historico_full


def processar_lancamentos_com_data_valor(lancamentos_por_data, formato='3col', ano_referencia=None):
    """
    Processa os lançamentos extraindo data, valor, tipo, cnpj/cpf, documento e pagador/recebedor
    """
    lancamentos_processados = []
    saldo_anterior = None
    ultimo_saldo_dia = None
    ultimo_saldo_do_dia_valido = None
    
    # Filtrar datas válidas antes de ordenar
    datas_validas = []
    for data in lancamentos_por_data.keys():
        try:
            # Verificar se a data é válida
            dia, mes = data.split('/')
            if 1 <= int(dia) <= 31 and 1 <= int(mes) <= 12:
                datas_validas.append(data)
        except (ValueError, IndexError):
            # Pular datas inválidas
            continue
    
    datas_ordenadas = sorted(datas_validas, 
                           key=lambda x: datetime.strptime(x, '%d/%m'))
    
    for data in datas_ordenadas:
        lancamentos = lancamentos_por_data[data]
        for lancamento in lancamentos:
            data_processada, valor_processado, tipo_operacao = extrair_data_valor(
                lancamento,
                formato,
                ano_referencia,
            )
            cnpj_cpf = extrair_cnpj_cpf(lancamento)
            if formato == '4col':
                documento, historico_4col = extrair_documento_historico_4col(lancamento)
                if not documento:
                    documento = extrair_documento(lancamento)
            else:
                documento = extrair_documento(lancamento)
                historico_4col = None
            pagador_recebedor = extrair_pagador_recebedor(lancamento, cnpj_cpf)
            
            # Verificar se é saldo anterior
            if "SALDO ANTERIOR" in lancamento and valor_processado is not None:
                saldo_anterior = abs(valor_processado)
                continue
            
            # Registrar saldos do dia úteis (ignorar bloco de resumo da última página)
            if "SALDO DO DIA" in lancamento and valor_processado is not None:
                if "RESUMO" not in lancamento.upper():
                    ultimo_saldo_do_dia_valido = abs(valor_processado)
                continue
            
            # Verificar se é o saldo final real (que aparece no resumo)
            if "SALDO EM C.CORRENTE" in lancamento or "SALDO DISPONÍVEL" in lancamento:
                saldo_final = extrair_saldo_final_resumo(lancamento)
                if saldo_final is not None:
                    ultimo_saldo_dia = saldo_final
                    continue
            
            if (data_processada and 
                valor_processado is not None and 
                tipo_operacao is not None and
                valor_processado != 0 and
                not eh_descricao_saldo(lancamento)):
                item = {
                    'data': data_processada,
                    'valor': valor_processado,
                    'tipo': tipo_operacao,
                    'cnpj_cpf': cnpj_cpf,
                    'documento': documento,
                    'pagador_recebedor': pagador_recebedor,
                    'lancamento_completo': lancamento
                }
                if formato == '4col':
                    item['historico_4col'] = historico_4col
                lancamentos_processados.append(item)
    
    if ultimo_saldo_dia is None and ultimo_saldo_do_dia_valido is not None:
        ultimo_saldo_dia = ultimo_saldo_do_dia_valido

    return lancamentos_processados, saldo_anterior, ultimo_saldo_dia

def exibir_lancamentos_processados(lancamentos_processados, saldo_anterior, ultimo_saldo_dia, debug=False):
    """
    Exibe os lançamentos processados de forma organizada
    
    Args:
        lancamentos_processados (list): Lista de lançamentos processados
        saldo_anterior (float): Saldo anterior do período
        ultimo_saldo_dia (float): Último saldo do dia
        debug (bool): Se True, exibe a lista detalhada de lançamentos
    """
    print("\n" + "="*80)
    print("CONFERÊNCIA DE SALDOS")
    print("="*80)
    
    # Exibir informações de saldo para conferência
    if saldo_anterior is not None:
        print(f"💰 Saldo Anterior do Período: R$ {saldo_anterior:10.2f}")
    if ultimo_saldo_dia is not None:
        print(f"💰 Saldo Final do Extrato: R$ {ultimo_saldo_dia:10.2f}")
    
    # Calcular saldo das movimentações, entradas e saídas
    if lancamentos_processados:
        saldo_movimentacoes = sum(l['valor'] for l in lancamentos_processados)
        total_entradas = sum(l['valor'] for l in lancamentos_processados if l['valor'] > 0)
        total_saidas = sum(l['valor'] for l in lancamentos_processados if l['valor'] < 0)
        print(f"💳 Saldo das Movimentações: R$ {saldo_movimentacoes:10.2f}")
        print(f"⬆️  Total de Entradas:      R$ {total_entradas:10.2f}")
        print(f"⬇️  Total de Saídas:        R$ {total_saidas:10.2f}")
        
        # Verificar se bate com o saldo do dia
        if saldo_anterior is not None and ultimo_saldo_dia is not None:
            saldo_calculado = saldo_anterior + saldo_movimentacoes
            print(f"🧮 Saldo Calculado (Anterior + Movimentações): R$ {saldo_calculado:10.2f}")
            diferenca = saldo_calculado - ultimo_saldo_dia
            print(f"🔍 Diferença: R$ {diferenca:10.2f}")
            if abs(diferenca) < 0.01:
                print("✅ CONFERÊNCIA: Saldos batem!")
            else:
                print("❌ CONFERÊNCIA: Saldos não batem!")
    print("-" * 80)
    
    # Exibir lista detalhada apenas em modo debug
    if debug:
        print("\n" + "="*80)
        print("LANÇAMENTOS COM DATA E VALOR PROCESSADOS")
        print("="*80)
        
        if not lancamentos_processados:
            print("❌ Nenhum lançamento válido encontrado!")
            return
        
        # Ordenar lançamentos cronologicamente
        lancamentos_ordenados = sorted(lancamentos_processados, key=lambda x: datetime.strptime(x['data'], '%d/%m/%Y'))
        
        print(f"📊 Total de lançamentos: {len(lancamentos_ordenados)}")
        print("-" * 80)
        print(f"{'#':>3}  {'Data':<10}  {'Valor':>12}  {'Tipo':<4}  {'CNPJ/CPF':<20}  {'Documento':<15}  {'Pagador/Recebedor':<20}  Lançamento")
        print("-" * 80)
        for i, lancamento in enumerate(lancamentos_ordenados, 1):
            data = lancamento['data']
            valor = lancamento['valor']
            tipo = lancamento['tipo']
            cnpj_cpf = lancamento.get('cnpj_cpf', '')
            documento = lancamento.get('documento', '')
            pagador_recebedor = lancamento.get('pagador_recebedor', '')
            texto = lancamento['lancamento_completo']
            # Formatar valor com sinal
            sinal = "+" if valor > 0 else ""
            valor_formatado = f"{sinal}{valor:.2f}"
            print(f"{i:3d}. {data:<10}  {valor_formatado:>12}  {tipo:<4}  {cnpj_cpf:<20}  {documento:<15}  {pagador_recebedor:<20}  {texto}")
        print("-" * 80)

def normalizar_historico_sicoob(lancamento_completo, pagador_recebedor, documento, formato='3col', historico_4col=None):
    """
    Gera histórico limpo para amarração.
    4col: usa históricocoluna (DÉB.IOF) separada do documento (IOF/2-1) — conforme layout da imagem
    3col: descrição + tipo Pix + nome + DOC
    """
    if formato == '4col' and historico_4col:
        # historico_4col já contém a descrição completa (incl. linhas de continuação)
        return historico_4col.strip()

    # 3col ou fallback
    partes = [p.strip() for p in lancamento_completo.strip().split('|') if p.strip()]
    primeira = partes[0] if partes else ''
    primeira = re.sub(r'^\d{2}/\d{2}(/\d{4})?\s*', '', primeira)
    primeira = re.sub(r'\s*-?\d{1,3}(?:\.\d{3})*,\d{2}\s*[DC]?\s*$', '', primeira).strip()
    tipo_pix = next((p for p in partes[1:] if p in ('Pagamento Pix', 'Recebimento Pix')), '')
    pedacos = [primeira, tipo_pix, pagador_recebedor.strip()] if pagador_recebedor else [primeira, tipo_pix]
    historico = ' '.join(p for p in pedacos if p).strip()
    if documento:
        historico += f" DOC: {documento}"
    return historico or lancamento_completo


def extrair_documento(texto):
    """
    Extrai o documento que aparece após "DOC.:" em um texto.
    Retorna string vazia se não encontrar.
    """
    import re
    
    # Padrão para encontrar "DOC.:" seguido do documento
    padrao = r'DOC\.:\s*([^|]+)'
    
    match = re.search(padrao, texto)
    if match:
        return match.group(1).strip()
    
    return ''

def extrair_pagador_recebedor(texto, cnpj_cpf):
    """
    Extrai o pagador/recebedor do texto.
    Se encontrar "Pagamento Pix", retorna o CNPJ/CPF.
    Caso contrário, busca na segunda, terceira ou quarta posição.
    """
    import re
    
    # Dividir a linha pelos separadores |
    partes = texto.split('|')
    
    # Se for "Pagamento Pix", retorna o CNPJ/CPF
    if 'Pagamento Pix' in texto:
        return cnpj_cpf
    
    # Buscar na segunda, terceira ou quarta posição
    for i in range(1, min(5, len(partes))):
        parte = partes[i].strip()
        
        # Pular se for apenas o tipo (C ou D)
        if parte in ['C', 'D']:
            continue
            
        # Pular se for "Pagamento Pix" ou "Recebimento Pix"
        if parte in ['Pagamento Pix', 'Recebimento Pix']:
            continue
            
        # Pular se for um CNPJ/CPF (já temos essa informação)
        if re.search(r'(?:\*{3}|\d{3})\.\d{3}\.\d{3}-(?:\d{2,3}|\*{2})', parte):
            continue
        if re.search(r'\d{2}\.\d{3}\.\d{3}[ /]\d{4}-\d{2}', parte):
            continue
            
        # Pular se for "DOC.:" ou começar com "DOC.:"
        if parte.startswith('DOC.:'):
            continue
            
        # Se chegou aqui, é provavelmente o pagador/recebedor
        if parte and len(parte) > 2:  # Evitar partes muito pequenas
            return parte
    
    return ''

def gerar_csv_simplificado(lancamentos_processados, caminho_pdf):
    """
    Gera um arquivo CSV (.csv) com apenas as colunas: data, nome, valor
    O nome inclui o sufixo "- SICOOB"
    """
    if not lancamentos_processados:
        print("❌ Nenhum lançamento para gerar CSV!")
        return
    
    # Ordenar lançamentos cronologicamente
    lancamentos_ordenados = sorted(lancamentos_processados, key=lambda x: datetime.strptime(x['data'], '%d/%m/%Y'))
    
    # Obter diretório e nome do arquivo PDF
    pdf_path = Path(caminho_pdf)
    diretorio = pdf_path.parent
    nome_base = pdf_path.stem
    
    # Nome do arquivo CSV
    nome_csv = f"{nome_base}_EXTRATO_SIMPLIFICADO.csv"
    caminho_csv = diretorio / nome_csv
    
    # Criar dados para CSV
    dados_csv = []
    for lancamento in lancamentos_ordenados:
        nome = lancamento.get('pagador_recebedor', '').strip()
        cnpj_cpf = lancamento.get('cnpj_cpf', '').strip()
        documento = lancamento.get('documento', '').strip()
        
        # Se nome e CNPJ/CPF estiverem em branco, usar o documento
        if not nome and not cnpj_cpf and documento:
            nome_com_sufixo = f"{documento} - SICOOB"
        elif nome and cnpj_cpf:
            # Verificar se o nome já contém o CNPJ/CPF para evitar duplicação
            if cnpj_cpf in nome:
                nome_com_sufixo = f"{nome} - SICOOB"
            else:
                nome_com_sufixo = f"{nome} {cnpj_cpf} - SICOOB"
        elif nome:
            nome_com_sufixo = f"{nome} - SICOOB"
        elif cnpj_cpf:
            nome_com_sufixo = f"{cnpj_cpf} - SICOOB"
        else:
            nome_com_sufixo = "SICOOB"
        
        # Formatar valor no padrão brasileiro (vírgula como separador decimal)
        valor_br = f"{lancamento['valor']:.2f}".replace('.', ',')
        
        dados_csv.append({
            'data': lancamento['data'],
            'nome': nome_com_sufixo,
            'valor': valor_br
        })
    
    # Criar DataFrame e salvar CSV com separador ponto e vírgula
    df_csv = pd.DataFrame(dados_csv)
    df_csv.to_csv(caminho_csv, index=False, encoding='utf-8-sig', sep=';')
    
    print(f"\n📄 CSV GERADO COM SUCESSO!")
    print(f"📁 Arquivo: {caminho_csv}")
    print(f"📈 Total de lançamentos: {len(dados_csv)}")
    
    return str(caminho_csv)

def extrair_texto_e_organizar(caminho_pdf, debug=False):
    """
    Extrai texto do PDF e organiza os lançamentos por data
    
    Args:
        caminho_pdf (str): Caminho para o arquivo PDF
        debug (bool): Se True, exibe informações detalhadas
    """
    texto, linhas = extrair_texto_pdf(caminho_pdf, debug)

    if texto:
        if debug:
            print("\n" + "=" * 50)
            print("EXTRAÇÃO CONCLUÍDA!")
            print(f"Total de caracteres extraídos: {len(texto)}")
        formato = detectar_formato_sicoob(linhas)
        if debug:
            print(f"Formato detectado: {formato} colunas")
        lancamentos_por_data = organizar_lancamentos_por_data(texto)
        exibir_lancamentos_organizados(lancamentos_por_data, debug)
        lancamentos_processados, saldo_anterior, ultimo_saldo_dia = processar_lancamentos_com_data_valor(lancamentos_por_data, formato)
        
        # Exibir lançamentos processados
        exibir_lancamentos_processados(lancamentos_processados, saldo_anterior, ultimo_saldo_dia, debug)
        
        # Gerar planilha Excel
        gerar_planilha_excel(lancamentos_processados, saldo_anterior, ultimo_saldo_dia, caminho_pdf)
        
        # Gerar CSV simplificado
        gerar_csv_simplificado(lancamentos_processados, caminho_pdf)

def gerar_planilha_excel(lancamentos_processados, saldo_anterior, ultimo_saldo_dia, caminho_pdf):
    """
    Gera um arquivo Excel (.xlsx) com os lançamentos processados
    """
    if not lancamentos_processados:
        print("❌ Nenhum lançamento para gerar planilha!")
        return
    
    # Ordenar lançamentos cronologicamente
    lancamentos_ordenados = sorted(lancamentos_processados, key=lambda x: datetime.strptime(x['data'], '%d/%m/%Y'))
    
    # Criar DataFrame
    dados = []
    for lancamento in lancamentos_ordenados:
        dados.append({
            'Data': lancamento['data'],
            'Valor': lancamento['valor'],
            'Tipo': lancamento['tipo'],
            'CNPJ/CPF': lancamento.get('cnpj_cpf', ''),
            'Documento': lancamento.get('documento', ''),
            'Pagador/Recebedor': lancamento.get('pagador_recebedor', ''),
            'Lançamento Completo': lancamento['lancamento_completo']
        })
    
    df = pd.DataFrame(dados)
    
    # Calcular totais
    total_entradas = sum(l['valor'] for l in lancamentos_ordenados if l['valor'] > 0)
    total_saidas = sum(l['valor'] for l in lancamentos_ordenados if l['valor'] < 0)
    saldo_movimentacoes = total_entradas + total_saidas
    
    # Obter diretório e nome do arquivo PDF
    pdf_path = Path(caminho_pdf)
    diretorio = pdf_path.parent
    nome_base = pdf_path.stem
    
    # Nome do arquivo Excel
    nome_excel = f"{nome_base}_EXTRATO_PROCESSADO.xlsx"
    caminho_excel = diretorio / nome_excel
    
    # Criar workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Extrato Processado"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Adicionar título
    ws['A1'] = f"EXTRATO BANCÁRIO - {nome_base}"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:G1')
    
    # Adicionar informações de saldo
    ws['A3'] = "RESUMO DE SALDOS"
    ws['A3'].font = Font(bold=True, size=12)
    ws.merge_cells('A3:G3')
    
    if saldo_anterior is not None:
        ws['A4'] = f"Saldo Anterior: R$ {saldo_anterior:,.2f}"
    if ultimo_saldo_dia is not None:
        ws['A5'] = f"Saldo Final: R$ {ultimo_saldo_dia:,.2f}"
    
    ws['A6'] = f"Total Entradas: R$ {total_entradas:,.2f}"
    ws['A7'] = f"Total Saídas: R$ {total_saidas:,.2f}"
    ws['A8'] = f"Saldo Movimentações: R$ {saldo_movimentacoes:,.2f}"
    
    # Adicionar cabeçalhos da tabela
    headers = ['Data', 'Valor', 'Tipo', 'CNPJ/CPF', 'Documento', 'Pagador/Recebedor', 'Lançamento Completo']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=10, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Adicionar dados
    for row_idx, lancamento in enumerate(lancamentos_ordenados, 11):
        # Data
        ws.cell(row=row_idx, column=1, value=lancamento['data']).border = border
        
        # Valor
        valor_cell = ws.cell(row=row_idx, column=2, value=lancamento['valor'])
        valor_cell.number_format = 'R$ #,##0.00'
        valor_cell.border = border
        
        # Tipo
        ws.cell(row=row_idx, column=3, value=lancamento['tipo']).border = border
        
        # CNPJ/CPF
        ws.cell(row=row_idx, column=4, value=lancamento.get('cnpj_cpf', '')).border = border
        
        # Documento
        ws.cell(row=row_idx, column=5, value=lancamento.get('documento', '')).border = border
        
        # Pagador/Recebedor
        ws.cell(row=row_idx, column=6, value=lancamento.get('pagador_recebedor', '')).border = border
        
        # Lançamento Completo
        ws.cell(row=row_idx, column=7, value=lancamento['lancamento_completo']).border = border
    
    # Ajustar largura das colunas
    ws.column_dimensions['A'].width = 12  # Data
    ws.column_dimensions['B'].width = 15  # Valor
    ws.column_dimensions['C'].width = 8   # Tipo
    ws.column_dimensions['D'].width = 25  # CNPJ/CPF
    ws.column_dimensions['E'].width = 15  # Documento
    ws.column_dimensions['F'].width = 30  # Pagador/Recebedor
    ws.column_dimensions['G'].width = 80  # Lançamento Completo
    
    # Salvar arquivo
    wb.save(caminho_excel)
    
    print(f"\n📊 PLANILHA GERADA COM SUCESSO!")
    print(f"📁 Arquivo: {caminho_excel}")
    print(f"📈 Total de lançamentos: {len(lancamentos_ordenados)}")
    
    return str(caminho_excel)

def main():
    """Função principal"""
    print("=== EXTRATOR DE TEXTO PDF ===\n")
    
    # Padronização: python conversor_extrato_sicoob_pdf_csv.py <entrada.pdf> <saida.csv> [conta_banco]
    if len(sys.argv) < 3:
        print("Uso: python conversor_extrato_sicoob_pdf_csv.py <arquivo.pdf> <arquivo_saida.csv> [conta_banco]")
        sys.exit(1)
    
    caminho_pdf = sys.argv[1]
    caminho_csv = sys.argv[2]
    conta_banco = sys.argv[3] if len(sys.argv) > 3 else '1.1.1.01'  # Conta padrão se não fornecida
    debug = "--debug" in sys.argv
    if not os.path.exists(caminho_pdf):
        print(f"Erro: O arquivo '{caminho_pdf}' não existe.")
        sys.exit(1)
    if not caminho_pdf.lower().endswith('.pdf'):
        print("Erro: O arquivo deve ser um PDF (.pdf)")
        sys.exit(1)
    texto, linhas = extrair_texto_pdf(caminho_pdf, debug)
    if texto:
        formato = detectar_formato_sicoob(linhas)
        if debug:
            print(f"Formato detectado: {formato} colunas")
        lancamentos_por_data = organizar_lancamentos_por_data(texto)
        lancamentos_processados, saldo_anterior, ultimo_saldo_dia = processar_lancamentos_com_data_valor(lancamentos_por_data, formato)
        # Padronizar saída para o formato esperado pelo importador avançado
        def formatar_valor_brl(valor):
            try:
                return f"{float(valor):,.2f}".replace('.', 'X').replace(',', '.').replace('X', ',')
            except Exception:
                return "0,00"
        
        with open(caminho_csv, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile, delimiter=';')
            # Cabeçalho com todas as colunas esperadas pelo importador avançado
            writer.writerow([
                'Data do Lançamento',
                'Usuário',
                'Conta Débito',
                'Conta Crédito', 
                'Valor do Lançamento',
                'Histórico',
                'Código da Filial/Matriz',
                'Nome da Empresa',
                'Número da Nota',
                'CNPJ/CPF',
            ])
            
            for l in lancamentos_processados:
                data = l['data']
                valor = l['valor']
                nome = l.get('pagador_recebedor', '')
                documento = l.get('documento', '')
                cnpj_cpf = l.get('cnpj_cpf', '')
                lancamento_completo = l.get('lancamento_completo', '').strip()
                historico_4col = l.get('historico_4col', '') if formato == '4col' else None

                # Histórico: 4col usa coluna HISTÓRICO isolada (ex: DÉB.IOF); 3col monta descrição
                historico = normalizar_historico_sicoob(lancamento_completo, nome, documento, formato, historico_4col)

                if valor > 0:
                    conta_debito = conta_banco
                    conta_credito = ''
                else:
                    conta_debito = ''
                    conta_credito = conta_banco

                writer.writerow([
                    data,
                    'Sistema',
                    conta_debito,
                    conta_credito,
                    formatar_valor_brl(valor),
                    historico,
                    '',
                    nome,
                    documento,
                    cnpj_cpf,
                ])
        print(f"CSV padronizado gerado em: {caminho_csv}")

if __name__ == "__main__":
    main()
